#!/usr/bin/env python 
# -*- coding:utf-8 -*-
# !@Author:liyihui
from openpyxl import Workbook, load_workbook

wb = Workbook()
# 活动sheet
ws = wb.active
# insert at the end (default)
ws1 = wb.create_sheet("Mysheet")
# insert at first position
ws2 = wb.create_sheet("Mysheet", 0)
# 倒数第二个位置插入
ws3 = wb.create_sheet("Mysheet", -1)
# 设置标题
ws.title = "New Title"
# 设置背景颜色
ws.sheet_properties.tabColor = "1072BA"
ws4 = wb["New Title"]
# 查看所有表名称
print(wb.sheetnames)
source = wb.active
# 创建副本
target = wb.copy_worksheet(source)
# 在内存中创建工作表时，它不包含任何单元格。它们是在首次访问时创建的。
c = ws['A4']
ws['A4'] = 4
d = ws.cell(row=2, column=4, value=10)
for x in range(1, 10):
    for y in range(1, 10):
        ws.cell(row=x, column=y)
cell_range = ws['A1':'C2']
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)
# 遍历所有行
ws['C9'] = 'hello world'
print(tuple(ws.rows))
# 遍历所有列
print(tuple(ws.columns))
# 遍历所有行中单元格值
for row in ws.values:
    for value in row:
        print(value)
for row in ws.iter_rows(min_row=1, max_col=10, max_row=10, values_only=True):
    print(row)
c.value = 'hello world'
print(c.value)
d.value = 3.14
print(d.value)
# 会覆盖现有文件，不发出警告
wb.save('balances.xlsx')
# 打开现有文件
wb2 = load_workbook('balances.xlsx')
print(wb2.sheetnames)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename=dest_filename)
