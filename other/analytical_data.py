#!/usr/bin/env python
# -*- coding: utf-8 -*-

# 安装python
# 新建一个文件夹
# 将附件1.xlsx（请勿改名）复制到该文件夹中，保证附件1与脚本属于同个文件夹中
# 按住shift键右键打开在【此处打开 Powershell窗口】
# 输入pip install openpyxl，安装openpyxl库
# 运行该脚本，耐心等待，所生成文件将在文件夹中出现
# 运行脚本时，请先关闭所有excel文件，重复执行脚本，生成的文件将直接被覆盖
from openpyxl import Workbook, load_workbook


def main():
    new_excel = Workbook()
    new_sheet = new_excel.active
    new_sheet.title = "Sheet0"
    # 添加首列标题
    new_sheet.append(["证券代码",
                      "证券名称",
                      "交易时间",
                      "开盘价",
                      "最高价",
                      "最低价",
                      "收盘价",
                      "涨跌",
                      "涨跌幅%",
                      "成交量",
                      "成交额",
                      "盘后量",
                      "盘后额"])
    # 设置交易时间的列宽，防止显示#####现象
    new_sheet.column_dimensions["C"].width = 20
    print("Reading data...Please wait！")
    # 读取excel文件数据
    original_table = load_workbook(filename="附件1.xlsx")
    # 获取所有sheet表名字
    sheets_name = original_table.sheetnames
    # 首列标题已添加，从第二行开始写数据
    crow = 2
    for i in range(len(sheets_name)):
        print("Analyze {0}".format(sheets_name[i]))
        ws = original_table[sheets_name[i]]
        # 获取该sheet的最大行数
        max_row = ws.max_row
        # 获取该sheet的最大列数
        max_column = ws.max_column
        # 跳过首列标题，从第二行开始读取数据，删除最后空了的五行
        data_list = []
        column_list = []
        for row in range(2, max_row-5):
            for col in range(1, max_column+1):
                data = ws.cell(row=row, column=col).value
                # 当表数据为【--】时使其等于0，防止后续制表时出错
                if data == "——":
                    data = 0
                column_list.append(data)
                new_sheet.cell(row=crow, column=col).value = data
            data_list.append(column_list)
            crow += 1
    original_table.close()
    print("Saving...")
    new_excel.save("StockData.xlsx")
    print("Complete!")


if __name__ == '__main__':
    main()
